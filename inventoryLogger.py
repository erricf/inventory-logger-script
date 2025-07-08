# Import the load_workbook function from the openpyxl library
# This is used to read Excel files (.xlsx)
from openpyxl import load_workbook

# Define a function that checks inventory levels in an Excel file
def check_inventory_excel(file_path):
    # Load the Excel workbook (spreadsheet) from the given file path
    wb = load_workbook(filename=file_path)
    
    # Select the active sheet (usually the first one that's open by default)
    sheet = wb.active

    # Go through each row in the sheet, starting from the second row (skipping headers)
    # 'values_only=True' means we only care about the actual data, not the cell formatting
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Unpack the values from each row into three variables
        product, stock, min_stock = row

        # Convert stock and minimum stock to integers so we can compare them
        # (If they are not already numbers, this could cause an error)
        stock = int(stock)
        min_stock = int(min_stock)

        # Print out the product and its stock information
        print(f"Product: {product}, Stock: {stock}, Minimum Stock: {min_stock}")
        
        # Check if the stock is below the minimum required level
        if stock < min_stock:
            # If it is, print an alert message
            print(f"--> ALERT: Restock needed for {product}!\n")

if __name__ == "__main__":
    check_inventory_excel('inventory.xlsx')